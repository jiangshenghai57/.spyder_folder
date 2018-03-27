# -*- coding: utf-8 -*-
"""
Created on Fri Feb  2 11:14:22 2018

Data scrubbing for Palides deals
Particular for my one deal Normandy 15-1
with Multi-servicer

@author: Shenghai

#Box
#########################################################################################################
#                                                                                                       #
#########################################################################################################
"""

import pandas as pd
import os
import sys
from shutil import copyfile
from openpyxl import load_workbook
import math
import numpy as np

global letters
letters = ["A",
        "B",
        "C",
        "D",
        "E",
        "F",
        "G",
        "H",
        "I",
        "J",
        "K",
        "L",
        "M",
        "N",
        "O",
        "P",
        "Q",
        "R",
        "S",
        "T",
        "U",
        "V",
        "W",
        "X",
        "Y",
        "Z",]

#########################################################################################################
# If variable empty return true else false                                                              #
#########################################################################################################
def is_empty(any_structure):
    if any_structure:
        return False
    else:
        return True
    

#########################################################################################################
# Asking user for month and year input                                                                  #
# This chunck of code will give the string info for what period of files to roll                        #
# More features could be add in to ask user                                                             #
#########################################################################################################
global dist_date
dist_date = "1802"
def user_input():
    print("What period would you like to roll the payment for Normandy \n" + 
          "If you want a specific month and year, \n")
    dist_date = input("Please enter the distrituion month and year, (example 1702 for 2017 February): ")
    return dist_date


#########################################################################################################
# Taking distribution date as an argument and return last month in 4-digit                              #
#########################################################################################################
def last_mon(dist_date):
    if int(dist_date[-2:]) == 1:
        last_mon = 12
        cur_yr = int(dist_date[0:2])
        last_yr = cur_yr - 1
        last_mon = str(last_yr) + str(last_mon)
    else:
        if (int(dist_date[-2:]) - 1) < 10:
            cur_mon = int(dist_date[-2:])
            last_mon = cur_mon - 1
            cur_yr = int(dist_date[0:2])
            last_mon = str(cur_yr) + "0" + str(last_mon)
        else:
            cur_mon = int(dist_date[-2:])
            last_mon = cur_mon - 1
            cur_yr = int(dist_date[0:2])
            last_mon = str(cur_yr) + str(last_mon)
             
    return last_mon


#########################################################################################################
# Copy the file from last month                                                                         #
#########################################################################################################
def copy_temp_rm():
    # Use this input later for final production
#    dist_date = "1702" # user_input()
    
    # Clean this up later for user input and excel file naming convention
    rm = "D:/deals/Carrington/SMAC/2015-01-N/remit/" + last_mon(dist_date) + "/ETI_TEMPLATE_RUSHMORE_" + last_mon(last_mon(dist_date)) + ".xlsx"

    # Read last month's end_bal and acct num
    global acct_num, beg_bal
    # First worksheet is 'RUSHMORE' raw data 
    # Second worksheet is remittance report data
    df = pd.read_excel(rm, sheetname = 1, header = 0)
    col_index = list(df.columns)
    
    df_last_month = pd.DataFrame(df[:][col_index[0]])
    df_last_month = df_last_month.join(df[:][col_index[1]])
    
    # Search for ending def_bal from last month and used it as beg_def_bal for current month
    for i in range(15, len(col_index)):
        if 'Ending Def Prin' in col_index[i]:
            index_num = i
            
    # Adding ending def_bal column
    df_last_month = df_last_month.join(df[:][col_index[index_num]])
    
    
    # Making folder for the distribution month
    # This could use the user_input()'s value 
    if not os.path.exists("D:/deals/Carrington/SMAC/2015-01-N/remit/" + dist_date):
        os.makedirs("D:/deals/Carrington/SMAC/2015-01-N/remit/" + dist_date)
    else:
        print("Directory already exists!!!")
        print("Seems like the file just got modified!!!")
        print("Be careful what you're doing!!!")
#        sys.exit()
    
    # Assign new name to this month's file name
    global cur_mon_file
    cur_mon_file = "D:/deals/Carrington/SMAC/2015-01-N/remit/" + dist_date + "/ETI_TEMPLATE_RUSHMORE_" + last_mon(dist_date) + ".xlsx"
    # Copy previous month file as a template for this month
    copyfile(rm, cur_mon_file)


#def read_remit_bsi():
    
#########################################################################################################
# This func will read monthly RUSHMORE REMIT file and clean up the data                                 #
# More features should be add to                                                                        #
# Spits out dataframe do not worry about FMV
#########################################################################################################
def read_remit_rm():
    global df_rm
    # Do a file search first to find the right investor num .xlsx extension
    cur_mon_dir = "D:/deals/Carrington/servicing_transfer/" + dist_date + "/RUSHMORE/"
    file_list = os.listdir(cur_mon_dir)
    
    for file in file_list:
        if 'Inv 228' in file and ".xls" in file:
            cur_mon_remit_rm = cur_mon_dir + file

    # Worksheet names for current month remit file RUSHMORE
    ######################
    #    NMLT 2015-1 REMIC      0
    #    Loan Level Recap       1
    #    Remittance Detail      2
    #    PIF                    3
    #    Trial Balance          4
    #    Escrow Advances        5
    #    Corp Advances          6
    #    3rd Party Advances     7
    #    Supplemental Funds     8
    #    Invoices               9
    #    Loan Sale Proceeds     10
    #    Loan Modification      12                                                                         
    ######################
    # Getting all the worksheets' names
    ws_names = pd.ExcelFile(cur_mon_remit_rm).sheet_names
    
    df_list = []
    
    # Search for the correct row to start and to be the columns index with Trial Balance worksheet
    for i in range(0, 20):
        df_tri_bal = pd.read_excel(cur_mon_remit_rm, sheetname = ws_names[4], header = i)
        col_index_tb = list(df_tri_bal.columns)
        if "loan number" in str(col_index_tb[0]).lower():
            break
    
    # Search for the correct column to grab the end_bal index
    for i in range(0, len(col_index_tb)):
        if 'first principal balance' in str(col_index_tb[i]).lower():
            end_bal_ind = i
            df_list.append(end_bal_ind)
        elif 'first p and i amount' in str(col_index_tb[i]).lower():
            pi_ind = i
            df_list.append(pi_ind)
        elif 'annual interest rate' in str(col_index_tb[i]).lower():
            air_ind = i
            df_list.append(air_ind)
        elif 'deferred principal balance' in str(col_index_tb[i]).lower():
            def_ind = i
            df_list.append(def_ind)
            
    # DataFrame RUSHMORE Trial Balance worksheet
    df_rm = pd.DataFrame(df_tri_bal[:][col_index_tb[0]])
    for i in range(0, len(df_list)):
        df_rm = df_rm.join(df_tri_bal[:][col_index_tb[df_list[i]]])
    
    df_list2 = []
    
    # DataFrame RUSHMORE Loan Level Recap worksheet
    for i in range(0, 20):
        df_llr = pd.read_excel(cur_mon_remit_rm, sheetname = ws_names[1], header = i)
        col_index_llr = list(df_llr.columns)
        if "loan" in str(col_index_llr[0]).lower() and "number" in str(col_index_llr[0]).lower():
            break
        
    # Search fo the corrrect columns to grab the svcr fee * -1, interest, and prin
    for i in range(0, len(col_index_llr)):
        if 'interest' in str(col_index_llr[i]).lower():
            end_bal_ind = i
            df_list2.append(end_bal_ind)
        elif 'principal' in str(col_index_llr[i]).lower():
            pi_ind = i
            df_list2.append(pi_ind)
        elif 'servicing fee' in str(col_index_llr[i]).lower():
            air_ind = i
            df_list2.append(air_ind) 
            
    # append df_list2 into df_rm (DataFrame RUSHMORE)
    for i in range(0, len(df_list2)):
        df_rm = df_rm.join(df_llr[:][col_index_llr[df_list2[i]]])
  


#########################################################################################################
# This function will put dataframe rushmore into the monthyly tempalte file                             #
#########################################################################################################
def create_raw_sh():
    # Load the current month RUSHMORE month file
    wb = load_workbook(cur_mon_file)
    ws = wb.create_sheet(last_mon(dist_date))
    sh_names = list(wb.get_sheet_names())
    col_index = list(df_rm.columns)
    row_index = list(df_rm.index)

    print(sh_names)

    for i in range(0, len(col_index)):
        ws[letters[i] + '1'] = col_index[i]
        for j in range(0, len(row_index)):
            ws[letters[i] + str(2 + j)] = df_rm[col_index[i]][row_index[j]]
    
    wb.save(cur_mon_file)
    
#    if is_empty(df):
#        print('You better reset the head setting in read_excel function')
    
#def read_remit_selene():    

if __name__ == '__main__':
    copy_temp_rm()
    read_remit_rm()
    create_raw_sh()
    sys.exit()