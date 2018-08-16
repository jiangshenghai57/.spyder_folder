# -*- coding: utf-8 -*-
"""
Created on Thu Mar  1 14:41:28 2018

@author: Shenghai
"""

#########################################################################################################
# This py script will scrub the monthly remit files for all CMS deals                                   #
# It will produce dat file and calculate the total net interest in the remit_scrub file
# User will only open up the remit file and copy and paste the row into the hist-job and other adj. if  #
# neccessary                                                                                            #
#########################################################################################################

import sys
import os
from shutil import copyfile
import xlwings as xw
from openpyxl import load_workbook
import numpy as np
from time import time

#########################################################################################################
# Prompt console for usper input. Follow print func below                                               #
#########################################################################################################
def user_input():
    # Asking user for month and year input
    # Retun the as a 4-digit output as year and month, 1701

    while True:
        dist_date = input("Please enter the distrituion month and year, (example 0117 for January 2017): ")
        try:
            if int(dist_date[0:2]) >= 1 and int(dist_date[0:2]) <= 12:
                month = int(dist_date[0:2])
                year = int(dist_date[-2:])
                if month < 10:
                    month = "0{}".format(str(month))
                else:
                    month = str(month)
                    
                dist_date = "{}{}".format(str(year), str(month))
                break
            else:
                continue
        except:
            pass

    # return dist_date as a string like 1801 for Jan. 2018
    return dist_date


#########################################################################################################
# If variable empty return true else false                                                              #
#########################################################################################################
def is_empty(any_structure):
    # empty structure is False by default
    if any_structure:
        return False
    else:
        return True


my_deals = {
                # My deals
                '2014-03': 'D:/deals/Carrington/SMAC/2014-03/',
                '2014-05': 'D:/deals/Carrington/SMAC/2014-05/',
                '2015-01': 'D:/deals/Carrington/SMAC/2015-01/',
                '2015-08': 'D:/deals/Carrington/SMAC/2015-08/',
                '2015-09': 'D:/deals/Carrington/SMAC/2015-09/',
                '2016-01': 'D:/deals/Carrington/SMAC/2016-01/',
                '2016-04': 'D:/deals/Carrington/SMAC/2016-04/',
                '2016-21': 'D:/deals/Carrington/SMAC/2016-21/',
                '2016-23': 'D:/deals/Carrington/SMAC/2016-23/',
                '2016-31': 'D:/deals/Carrington/SMAC/2016-31/',
                '2016-32': 'D:/deals/Carrington/SMAC/2016-32/',

                # Brian's deals
#                '2013-20': 'D:/deals/Carrington/SMAC/2013-20/',
#                '2013-21': 'D:/deals/Carrington/SMAC/2013-21/',
#                '2014-01': 'D:/deals/Carrington/SMAC/2014-01/',
#                '2014-04': 'D:/deals/Carrington/SMAC/2014-04/',
#                '2015-03': 'D:/deals/Carrington/SMAC/2015-03/',
#                '2015-04': 'D:/deals/Carrington/SMAC/2015-04/',
#                '2015-05': 'D:/deals/Carrington/SMAC/2015-05/',
#                '2015-07': 'D:/deals/Carrington/SMAC/2015-07/',
#                '2016-05': 'D:/deals/Carrington/SMAC/2016-05/',
#                '2016-06': 'D:/deals/Carrington/SMAC/2016-06/',
#                '2016-07': 'D:/deals/Carrington/SMAC/2016-07/',
#                '2016-08': 'D:/deals/Carrington/SMAC/2016-08/',
#                '2016-09': 'D:/deals/Carrington/SMAC/2016-09/',
#                '2016-10': 'D:/deals/Carrington/SMAC/2016-10/',
#                '2016-11': 'D:/deals/Carrington/SMAC/2016-11/',
#                '2016-12': 'D:/deals/Carrington/SMAC/2016-12/',
#                '2017-01': 'D:/deals/Carrington/SMAC/2017-01/',
#                '2017-02': 'D:/deals/Carrington/SMAC/2017-02/',
#                '2017-03': 'D:/deals/Carrington/SMAC/2017-03/',
#                '2017-04': 'D:/deals/Carrington/SMAC/2017-04/',
#                '2017-05': 'D:/deals/Carrington/SMAC/2017-05/',
#                '2017-06': 'D:/deals/Carrington/SMAC/2017-06/',
#                '2017-09': 'D:/deals/Carrington/SMAC/2017-09/',
#                '2017-21': 'D:/deals/Carrington/SMAC/2017-21/',
#                '2017-31': 'D:/deals/Carrington/SMAC/2017-31/',
#                '2017-32': 'D:/deals/Carrington/SMAC/2017-32/',
#                '2017-33': 'D:/deals/Carrington/SMAC/2017-33/',
#                '2017-34': 'D:/deals/Carrington/SMAC/2017-34/',
                }

#########################################################################################################
# Scrubbing the remit file                                                                              #
#########################################################################################################
#@profile
def scrub_carr():
    
    # Asking user to input a 4-digit year and month (e.g. 1812 for 2018 December)
    dist_date = user_input()
    # Dictionary of my carrington deals (Brian's deals)
    # So the for loop can go through each deal's remit files

    # Macro book is in D drive execs folder
    wb = xw.Book('d:/execs/ETI CMS.xlam')
    # Creating a python version of the macro, past_hist_row() takes two arguments
    # Will be used in the end of the deal for loop
    paste_hist_row = wb.macro('paste_hist_row')

    # This the main for loop to iterate throught each deal key
    range_deal = np.array(list(my_deals.keys()))

    for deal in range_deal:

        # Outter lvl exception handling if one deal does not go well,
        # it will go to the next deal
#        try:
            path = my_deals[deal]
            remit_path = "{}/remit/".format(path)

            # current monthly remit if there is more than 1 file in the dir,
            # make user choose which file to use
            cur_mon_remit = "{}/remit/{}".format(path, dist_date)

            # To see how many files is in the directory
            remit_files = np.array(os.listdir(cur_mon_remit))

            if not os.path.exists(cur_mon_remit):
                print('Unfortunately this {} directory does not exist.'.format(cur_mon_remit))
                print("Make sure the path name is correct.")

            # If there is no remit in the remit/dist_date/ folder continue to the next deal
            elif len(remit_files) == 0:
                print("There is no file in {}/remit/{}".format(path, dist_date))
                print("Going to the next deal")
                continue

            # This condition should be the common one after just copied over the remit files from Tom's share
            elif (len(remit_files) == 1) and (dist_date[0:2] in remit_files[0]) and (".xlsx" in remit_files[0]):
                cur_mon_remit = "{}/{}".format(cur_mon_remit, remit_files[0])

            # Most common one, this condition may be unneccessary
#            elif len(remit_files) == 1:
#                cur_mon_remit = "{}/{}".format(cur_mon_remit, remit_files[0])

            # If rerun again scrub will be overwritten
            elif len(remit_files) == 2 and "_scrub" not in remit_files[0] and '_scrub' in remit_files[1]:
                cur_mon_remit = "{}/{}".format(cur_mon_remit, remit_files[0])

            # If there is more than two files and other than remit and remit scrub than ask user to choose
            else:
                for file in remit_files:
                    print("{}. {}".format(str(remit_files.index(file)), file))

                print("Which file would you like to be your {} monthly remit file ".format(dist_date))

                while True:
                    file_num = input('Please Choose a File number: ')
                    try:
                        if int(file_num) < 0 and int(file_num) >= len(remit_files):
                            pass
                        else:
                            cur_mon_remit = '{}/{}'.format(cur_mon_remit,remit_files[int(file_num)])
                            print("This file '{}' is going to be your {} remit file".format(remit_files[int(file_num)], dist_date))
                            print('Processing...')
                            break
                    except:
                        break

            # Copy the current month file and rename, and keep the original remit
            if '.xlsx' in cur_mon_remit[-5:]:
                remit_scrub = list(cur_mon_remit)
                remit_scrub = ''.join(remit_scrub[0: (len(remit_scrub) - 5)]) + '_scrub.xlsx'
                copyfile(cur_mon_remit, remit_scrub)
            # Most likely it will not hit this condition. All remit files are in the current excel format xlsx
            elif '.xls' in cur_mon_remit[-4:]:
                print('Seems like the program cannot process the old excel format.')
                continue
            else:
                print('File cannot be found!!!')
                continue

            # Find out the remit files worksheet names
            # load_workbook will take a bit of time, since it's processing the entire remit file
            # There are other ways to do this to cut down processing time, but it's not a priority right now
            wb = load_workbook(remit_scrub, read_only=False)
            wrksht_nm = np.array(wb.sheetnames)

            # Ensure there is a Remittance Report Worksheet
            for name in wrksht_nm:
                if 'remittance report' in name.lower():
                    ws = wb[name]
                    break
                else:
                    ws = None
                    continue

            # Addition checking
            if not ws:
                print('Not sure which worksheet you want to look at.')
                print('You might want to check out the remit file and/or change the Worksheet\'s name to Remittance Report.')
                print('Going to the next deal.')
                continue
            else:
                pass

            # iteration list
            hun_k = np.arange(1, 100000)

            # Searching for loan count row and find the hist job row
            for i in hun_k:
                if "loan count" in str(ws['A' + str(i)].value).lower():
                    row = str(i)
                    break

            # Detecting more than one empty row and record them
            empty_row = []
            append = empty_row.append
            range1 = np.arange(1, int(row) - 1)
            for i in range1:
                if not ws['A' + str(i)].value:
                    append(str(i))

            # First checker to ensure the CMS remit file format is correct
            # Beginning Balance should be in the G column, otherwise it would go to the next deal
            if ("beginning" not in str(ws['G1'].value).lower()) or ('balance' not in str(ws['G1'].value).lower()):
                print('Seems like the file format has changed for {} distribution file.'.format(dist_date))
                print('You might want to check it out.')
                print('Going to the next deal.')
                continue

            # Hist job row. Will be used for checking first, then changed to plus calc int and adjust prin write, etc.
            hist_row = []
            append = hist_row.append
            hist_row_range = np.arange(6, 33)

            # This list may be unneccessary, library has a feature to get letters for column
            lttrs = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V',
             'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL', 'AM', 'AN', 'AO',
             'AP', 'AQ', 'AR', 'AS', 'AT', 'AU', 'AV', 'AW', 'AX', 'AY', 'AZ',]

            for i in hist_row_range:
                append(ws[lttrs[i] + row].value)

            # Checking if the total beg balnce with hist row
            tot_beg_bal = 0.0

            # If it encounters multiple empty_row
            # or there is no empty rows at all
            # decide to select where to stop
            # It does not delete the data in remit files, it just stops calculating at stop row
            # Additional features can be added in
            if not empty_row:
                # If there is nothing in the empty row, the row before hist_job_row is empty
                stop_row = str(int(row) - 1)
            elif len(empty_row) == 1:
                stop_row = empty_row[0]
            elif len(empty_row) == 2 and empty_row[1] == int(row) - 1:
                stop_row = empty_row[1]
            elif len(empty_row) == 2:
                # if there are only two empty rows, make sure empty_row[1] include all the inactive loans
                range3 = np.arange(2, int(empty_row[1]))
                for j in range3:
                    if is_empty(ws['G' + str(j)].value) == False:
                        tot_beg_bal = tot_beg_bal + ws['G' + str(j)].value

                    # Check if all the loans add up to beg_bal
                    if tot_beg_bal == hist_row[0]:
                        stop_row = int(empty_row[1])
                        break
                    elif (tot_beg_bal - hist_row[0] < 0.001):
                        stop_row = int(empty_row[1] )
                        break
                    else:
                        range4 = np.arange(2, int(empty_row[0]))
                        for j in range4:
                            if is_empty(ws['G' + str(j)].value) == False:
                                tot_beg_bal = tot_beg_bal + ws['G' + str(j)].value

                        # Check if all the loans add up to beg_bal
                        if tot_beg_bal == hist_row[0]:
                            stop_row = int(empty_row[0])
                            break
                        elif (tot_beg_bal - hist_row[0] < 0.001):
                            stop_row = int(empty_row[0] )
                            break
                        else:
                            print('Check out {} {} remit file!'.format(deal, dist_date))
                            print('You might want to delete extra empty rows.')
                            continue

            elif len(empty_row) == 3 and empty_row[-1] == int(row) - 2:
                # If the last empty row is right above the loan count row use the
                empty_row.pop()
                range5 = np.arange(0, len(empty_row))
                for i in range5:
                    for j in np.arange(2, int(empty_row[i])):
                        if is_empty(ws['G' + str(j)].value):
                            tot_beg_bal = tot_beg_bal + ws['G' + str(j)].value

                    # Check if all the loans add up to beg_bal
                    if tot_beg_bal == hist_row[0]:
                        stop_row = int(empty_row[i])
                        break
                    elif (tot_beg_bal - hist_row[0] < 0.001):
                        stop_row = int(empty_row[i] )
                        break
                    else:
                        continue

            elif len(empty_row) == 3 and int(empty_row[0]) == int(empty_row[1]) - 1:
                empty_row.pop()
                range6 = np.arange(0, len(empty_row))
                for i in range6:
                    for j in np.arange(2, int(empty_row[i])):
                        if is_empty(ws['G' + str(j)].value):
                            tot_beg_bal = tot_beg_bal + ws['G' + str(j)].value

                    # Check if all the loans add up to beg_bal
                    if tot_beg_bal == hist_row[0]:
                        stop_row = int(empty_row[i])
                        break
                    elif (tot_beg_bal - hist_row[0] < 0.001):
                        stop_row = int(empty_row[i] )
                        break
                    else:
                        continue

            elif len(empty_row) > 3:
                range7 = np.arange(0, len(empty_row))
                for i in range7:
                    for j in np.arange(2, int(empty_row[i])):
                        if is_empty(ws['G' + str(j)].value):
                            tot_beg_bal = tot_beg_bal + ws['G' + str(j)].value

                    # Check if all the loans add up to beg_bal
                    if tot_beg_bal == hist_row[0]:
                        stop_row = int(empty_row[i])
                        break
                    elif (tot_beg_bal - hist_row[0] < 0.001):
                        stop_row = int(empty_row[i] )
                        break
                    else:
                        continue
            else:
                print('Check out {} {} remit file!'.format(deal, dist_date))
                print('You might want to delete extra empty rows.')
                continue

            # check if AG is still fmv column
            # Additinal checker for CMS format
            if "fmv" in (ws['AG1'].value).lower():
                pass
            else:
                print('Format has changed!!!')
                print('Check the new FMV column in '.format(remit_scrub))
                continue

            # if there is a fmv, then erase def prin, prin write-off, new formula in hist_row
            # Sum up the transferred balance
            # Assign new formula into the hist row
            # This can be changed into for loops to input hard numbers onto the hist_row

            def_prin_w_off = 0
            prin_w_off = 0
            end_def_prin_w_off = 0

            loan_range = np.arange(2, int(stop_row))
            trans_bal = 0
            def_trans_bal = 0
            for num in loan_range:
                if (is_empty(ws['AG' + str(num)].value) == True):
                    continue
                elif ws['AG' + str(num)].value > 0:
                    ws['Q' + str(num)] = 0
                    ws['R' + str(num)] = 0
                    ws['AD' + str(num)] = 0
                    trans_bal = trans_bal + ws['G' + str(num)].value
                    def_trans_bal = def_trans_bal + ws['AC' + str(num)].value

            for num in loan_range:
                if is_empty(ws['Q' + str(num)].value) == False:
                    def_prin_w_off = def_prin_w_off + ws['Q' + str(num)].value
                else:
                    pass

                if is_empty(ws['R' + str(num)].value) == False:
                    prin_w_off = prin_w_off + ws['R' + str(num)].value
                else:
                    pass

                if is_empty(ws['AD' + str(num)].value) == False:
                    end_def_prin_w_off =  end_def_prin_w_off + ws['AD' + str(num)].value
                else:
                    pass

            ws['E' + str(row)] = trans_bal
            ws['D' + str(row)] = trans_bal + def_trans_bal
            ws['Q' + str(row)] = def_prin_w_off
            ws['R' + str(row)] = prin_w_off
            ws['AD' + str(row)] = end_def_prin_w_off

            hist_row2 = []

            for i in hist_row_range:
                hist_row2.append(ws[lttrs[i] + row].value)

            # Replace none with zero
            range8 = np.arange(0, len(hist_row2))
            for i in range8:
                if is_empty(hist_row2[i]) == True:
                    hist_row2[i] = 0

            # Create the dat file in the deal remit folder
            # 7-columns carrington dat
            # carrington #, beg_bal, end_bal, prin, gross int, prin w/r/loss, end def prin

            # Search for other dat and/or inp file to create deal name and date extention

            c_file = os.listdir(path)
            for file in c_file:
                if '.inp' in file:
                    deal_name = str(file[0:5])
                    break
                else:
                    deal_name = None
            
            '''
            if is_empty(deal_name) == False:
                dat = open('{}{}_{}.dat'.format(remit_path, deal_name, dist_date), 'w+')
            else:
                dat = open('{}_{}.dat'.format(remit_path, dist_date), 'w+')
                print('Make sure you rename the dat file')


            # Create the monthly distribution dat file (Carrington 7-column)
            # Wipe out the end_def_prin if there is a fmv
            # If fmv is greater than beg_bal, produce a negative write off in the dat file
            
            for loan in loan_range:
                # (1) Carrington Ln#
                if is_empty(ws['A' + str(loan)].value) == False:
                    dat.write("{} ".format(str(ws['A' + str(loan)].value)))
                else:
                    pass

                # (2) beg_bal, if not empty write beg bal
                if is_empty(ws['G' + str(loan)].value) == False:
                    dat.write('{0:.2f} '.format(ws['G' + str(loan)].value))
                else:
                    dat.write("0.00 ")

                # (3) end bal, if not empty write end bal
                if is_empty(ws['H' + str(loan)].value) == False:
                    dat.write('{0:.2f} '.format(ws['H' + str(loan)].value))
                else:
                    dat.write("0.00 ")

                # (4) prin payment, if not empty write n column
                #                   else if not empty in prin payment and there is a fmv, write prin_pmt + fmv
                #                   else if empty in prin_pmt and there is a fmv, write fmv
                #                   else write zero
                if (is_empty(ws['N' + str(loan)].value) == False):
                    dat.write('{0:.2f} '.format(ws['N' + str(loan)].value))
                elif (is_empty(ws['N' + str(loan)].value) == False) and (is_empty(ws['AG' + str(loan)].value) == False):
                    dat.write('{0:.2f} '.format(ws['N' + str(loan)].value + ws['AG' + str(loan)].value))
                elif (is_empty(ws['N' + str(loan)].value) == True) and (is_empty(ws['AG' + str(loan)].value) == False):
                    dat.write('{0:.2f} '.format(ws['AG' + str(loan)].value))
                else:
                    dat.write("0.00 ")

                # (5) int, if not empty write interest
                if is_empty(ws['K' + str(loan)].value) == False:
                    dat.write('{0:.2f} '.format(ws['K' + str(loan)].value))
                else:
                    dat.write("0.00 ")

                # (6) prin w/r/loss, if not empty write prin_w_off
                #                    else if beg_def_bal not empty and fmv not empty, beg_bal - fmv + beg_def_bal
                #                    else if
                if is_empty(ws['R' + str(loan)].value) == False:
                    dat.write('{0:.2f} '.format(ws['R' + str(loan)].value))
                elif (is_empty(ws['AC' + str(loan)].value) == False) and (is_empty(ws['AG' + str(loan)].value) == False):
                    dat.write('{0:.2f} '.format(ws['G' + str(loan)].value - ws['AG' + str(loan)].value + ws['AC' + str(loan)].value))
                elif (is_empty(ws['AG' + str(loan)].value) == False):
                    dat.write('{0:.2f} '.format(ws['G' + str(loan)].value - ws['AG' + str(loan)].value))
                elif is_empty(ws['AG' + str(loan)].value) == True and is_empty(ws['R' + str(loan)].value) == True and is_empty(ws['Q' + str(loan)].value) == False:
                    dat.write('{0:.2f} '.format(ws['R' + str(loan)].value))
                else:
                    dat.write("0.00 ")

                # (7) end def prin
                if is_empty(ws['AD' + str(loan)].value) == False:
                    dat.write('{0:.2f}\n'.format(ws['AD' + str(loan)].value))
                else:
                    dat.write("0.00\n")

            dat.close()
            print("Deal {} dat file created in the remit folder.".format(deal))
            '''

            # Create net calc int for each loan
            # Repalce total_upb cell with total net interest
            total_net_int = 0
            for num in loan_range:
                if is_empty(ws['G' + str(num)].value) == False:
                    total_net_int = total_net_int + ws['G' + str(num)].value * (ws['J' + str(num)].value - ws['M' + str(num)].value) / 1200.0

            ws['AH' + str(int(row) - 1)] = "total net interest"
            ws['AH' + row] = round(total_net_int, 2)

            wb.save(remit_scrub)
            print("Deal {} remit file scrubbed.".format(deal))

            # Find the hist job in the deal folder
            # If there is more than one hist job, let user to choose which hsit-job to use
            hist_job = []
            deal_files = list(os.listdir(path))
            for file in deal_files:
                if 'hist-job' in file.lower():
                    hist_job.append(file)

            if is_empty(hist_job) == True:
                print("We cannot find the HIST-JOB file for deal ".join(deal))
                break
            elif len(hist_job) == 1:
                hist_job_file = path + '/' + hist_job[0]
            else:
                for file in hist_job:
                    print("{}. {}".format(hist_job.index(file), file))

                while True:
                    file_num = input('Please Choose a File number: ')
                    try:
                        if int(file_num) < 0 or int(file_num) >= len(hist_job):
                            continue
                        else:
                            hist_job_file = path + '/' + hist_job[int(file_num)]
                            break
                    except:
                        break

            #concatenate hist_row2
            hist_row2.append(total_net_int)

            for num in hist_row2:
                hist_row2[hist_row2.index(num)] = round(float(num), 2)

            # Joing hist by comma
            hist_row2_string = ",".join(map(str, hist_row2))

            # Open up the hist-job in the background
            xw.Book(hist_job_file)

            # Use the macro delared earlier in the function
            paste_hist_row(dist_date, hist_row2_string)
            print("Deal {} hist row pasted.".format(deal))

            
#        except:
#            print("Tried to process " + deal + " deal, but you might want to check it out.")
#            input("Please press Enter to continue ...")
#            print('')
#            continue


if __name__ == '__main__':
    scrub_carr()
    input('Please press Enter to exit...')
    sys.exit()
