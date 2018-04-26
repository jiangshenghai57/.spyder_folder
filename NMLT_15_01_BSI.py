# -*- coding: utf-8 -*-
"""
Spyder Editor

This is a temporary script file.
This script file is attempting to clean up the multi-servicer deal(s).

#Box
#########################################################################################################
#                                                         #
#########################################################################################################
"""

import pandas as pd
import os
import sys
from shutil import copyfile
from openpyxl import load_workbook
import math
import numpy as np

#########################################################################################################
# If variable empty return true else false                                                              #
#########################################################################################################
def is_empty(any_structure):
    if any_structure:
        return False
    else:
        return True


#########################################################################################################
# Remove any nan value in an int/string array                                                           #
#########################################################################################################
def remove_nan(_list_):
    _list_ = [value for value in _list_ if not math.isnan(value)]
    return _list_


#########################################################################################################
# Asking user for month and year input                                                                  #
# This chunck of code will give the string info for what period of files to roll                        #
# More features could be add in to ask user                                                             #
#########################################################################################################
global dist_date
dist_date = "1803"
def user_input():
    print("What period would you like to roll the payment for Normandy \n" + 
          "If you want a specific month and year, \n")
    dist_date = input("Please enter the distrituion month and year, (example 1702 for 2017 February): ")
    return dist_date


#########################################################################################################
# Taking distribution date as an argument and return last month in 4-digit                              #
#########################################################################################################
def last_mon(dist_date):
    if int(dist_date[-2:]) == 1:
        last_mon = 12
        cur_yr = int(dist_date[0:2])
        last_yr = cur_yr - 1
        last_mon = str(last_yr) + str(last_mon)
    else:
        if (int(dist_date[-2:]) - 1) < 10:
            cur_mon = int(dist_date[-2:])
            last_mon = cur_mon - 1
            cur_yr = int(dist_date[0:2])
            last_mon = str(cur_yr) + "0" + str(last_mon)
        else:
            cur_mon = int(dist_date[-2:])
            last_mon = cur_mon - 1
            cur_yr = int(dist_date[0:2])
            last_mon = str(cur_yr) + str(last_mon)
             
    return last_mon


#########################################################################################################
# Taking distribution date as an argument and return next month in 4-digit                              #
#########################################################################################################
#def next_mon(dist_date):
#    if int(dist_date[-2:]) == 12:
#        next_mon = '01'
#        cur_yr = int(dist_date[0:2])
#        next_yr = cur_yr + 1
#        next_mon = str(next_yr) + next_mon
#    else:
#        if (int(dist_date[-2:]) + 1) < 10:
#            cur_mon = int(dist_date[-2:])
#            next_mon = cur_mon + 1
#            cur_yr = int(dist_date[0:2])
#            next_mon = str(cur_yr) + "0" + str(next_mon)
#            
#    return next_mon


#########################################################################################################
# Copy the file from last month                                                                         #
# This block will find the correct end_bal and account number to be used for current month              #
#########################################################################################################
def copy_temp():
    # Use this input later for final production
#    dist_date = "1702" # user_input()
    
    # Clean this up later for user input and excel file naming convention
    BSI = "D:/deals/Carrington/SMAC/2015-01-N/remit/" + last_mon(dist_date) + "/ETI_TEMPLATE_BSI_" + last_mon(last_mon(dist_date)) + ".xlsx"

    # Read last month's end_bal and acct num
    global acct_num, beg_bal
    df = pd.read_excel(BSI)
    col_ind = list(df.columns) # Assign column index into a workable array
    acct_num = list(df[col_ind[0]][:]) # Acct nums from last month
    beg_bal = list(df[col_ind[2]][:]) # End_bal from last month, beg_bal for this month
    
    # Making folder for the distribution month
    # This could use the user_input()'s value 
    if not os.path.exists("D:/deals/Carrington/SMAC/2015-01-N/remit/" + dist_date):
        os.makedirs("D:/deals/Carrington/SMAC/2015-01-N/remit/" + dist_date)
    else:
        print("Directory already exists!!!")
        print("Seems like the file just got modified.")
        print("Be careful what you're doing!!!")
#        sys.exit()
    
    # Assign new name to this month's file name
    global cur_mon_file
    cur_mon_file = "D:/deals/Carrington/SMAC/2015-01-N/remit/" + dist_date + "/ETI_TEMPLATE_BSI_" + last_mon(dist_date) + ".xlsx"
    # Copy previous month file as a template for this month
    copyfile(BSI, cur_mon_file)
    
    
#########################################################################################################
# Go into D:\deals\Carrington\ dir, grabbing distribution month remit file                              #
# All information needed is stored as a DataFrame structur                                              #
# There is a account number to refer to for ending balances, pi payment, etc
#########################################################################################################
def read_remit():
    # Do a file search first to find the right investor num and deal name and remic.xlsx extension
    cur_mon_dir = "D:/deals/Carrington/servicing_transfer/" + dist_date + "/BSI/"
    file_list = os.listdir(cur_mon_dir)
    for i in range(0, len(file_list)):
        if (("NMLT" in file_list[i]) and ("2015-1" in file_list[i]) and ("REMIC" in file_list[i])):
            cur_mon_remit = cur_mon_dir + file_list[i]
    
    # Worksheet names in monthly remit file
    #############################################################################################################
    # ['SUMMARY', 'COVERSHEET', 'EXPENSES', 'CORP ADVANCE AR', 'OT 16-16 (PRP) LOAN SALE', 'COLLECTION - Z510', # 
    # 'ADVANCE-AGG', 'ADVANCE RECOVERY', 'TRIAL BALANCE - Z305', 'DELINQUENCY - D016', 'LOANS ADDED - Z504',    #
    # 'LOANS REMOVED - Z120', 'LOSS MIT TEST', 'CORP ADVANCE DRM', 'CORP ADVANCE GLOBAL', 'WRITTEN OFF',        #
    # 'INSURANCE TRACKING DATA', 'COLLECTION-AGG', 'COLLECTION NEW', 'CLOSED PROCEEDS']                         #
    #############################################################################################################
    # In SUMMARY worksheet tab, skipping the Carrington image, use row 2 as the header index  
    ws_names = pd.ExcelFile(cur_mon_remit).sheet_names
    df = pd.read_excel(cur_mon_remit, sheetname = ws_names[0], header = 1)
    col_ind = list(df.columns)
    
    global special_cell, prev_svcr_fee, corp_rec, special_cell2
    
    # Search for 'Advance Recovery tab' string; it will be stored and used in BSI template remit report file
    # Another string search needs to be added for Loan Sale
    for i in range(0, len(df[col_ind[0]][:])):
        if "Advance Recovery tab" in str(df[col_ind[0]][i]):
            corp_rec = df[col_ind[2]][i]
        elif "Corp Advance Global tab" in str(df[col_ind[0]][i]):
            corp_adv = df[col_ind[2]][i]
        elif "BSI Invoicing Fees" in str(df[col_ind[0]][i]):
            inv_fee = df[col_ind[2]][i]
        elif "Remittance Funds Due" in str(df[col_ind[0]][i]):
            rem_due = - df[col_ind[2]][i]
        elif "Loan Sale" in str(df[col_ind[0]][i]):
            special_cell2 = - df[col_ind[2]][i]
            
    # In COVERSHEET tab looking for fees, store them and will be used in remit report 
    df2 = pd.read_excel(cur_mon_remit, sheetname = ws_names[1], header = 1)
    col_ind2 = list(df2.columns)
    
    # Search for 'External Vendor Fees' and 'Current Total Fees Due' to be used in the 'special!!!' cell in remit file
    for i in range(0, len(df2[col_ind2[0]][:])):
        if "External Vendor Fees" in str(df2[col_ind2[0]][i]):
            ex_ven_fee = df2[col_ind2[3]][i]
        elif "Current Total Fees Due" in str(df2[col_ind2[0]][i]):
            cur_tot_fee = df2[col_ind2[3]][i]        
        elif "Expense Tab" in str(df2[col_ind2[0]][i]):
            exp_tab = - df2[col_ind2[3]][i]

    if dist_date == "1712":
        corp_adv = 0
        
    special_cell = rem_due + ex_ven_fee + cur_tot_fee
    prev_svcr_fee = exp_tab + corp_adv - inv_fee + ex_ven_fee + cur_tot_fee + rem_due
#    print("Specil!!! cell value is " + str(round(special_cell, 2)))
#    print("Svcring fee-Prev " + str(round(prev_svcr_fee, 2)))

    ###########################################################################
    # Search for 'TRIAL BALANCE - Z305' worksheet                             #
    ###########################################################################
    for i in range(0, len(ws_names)):
        if "TRIAL BALANCE - Z305" in ws_names[i]:
            t_bal = i
    # Trial balance worksheet        
    df3 = pd.read_excel(cur_mon_remit, sheetname = ws_names[t_bal])
    col_ind3 = list(df3.columns)
    for i in range(0, len(col_ind3)):
        if "ACCOUNT_NUMBER" in col_ind3[i]:
            acct_num_ind = i
            df3 = pd.read_excel(cur_mon_remit, sheetname = ws_names[t_bal], index_col = acct_num_ind)
        else:
            pass
    
    # Global variables will be used in write template functions to put into excel
    # cur_acct_num should match up with what is on the template file for the current month
    global end_bal, pi_pmt, note_rate, svcr_fee, defer_bal, cur_acct_num, row_ind5
    global end_bal_index, pi_pmt_index, note_rate_index, svcr_fee_index, defer_bal_index
    
#    print(df3[]['INVESTOR_CODE'])
    # Current Acct Number for current rolling month, this will be used in the excel template file
    # Remove any NaN values using remove_nan func
    cur_acct_num = list(df3.index)
    cur_acct_num = remove_nan(cur_acct_num)

    # Ending balances for each account
    for i in range(0, len(col_ind3)):
        if "END_UPB" in col_ind3[i]:
            end_upb = i
    end_bal = pd.DataFrame(df3[col_ind3[end_upb]][cur_acct_num[:]], columns = ['END_UPB'])
    end_bal_index = list(end_bal.index)

    # Find pi_pmt in the Trial balance worksheet
    for i in range(0, len(col_ind3)):
        if "PI_PMT" in col_ind3[i]:
            pi_pmt_ind = i
    pi_pmt = pd.DataFrame(df3[col_ind3[pi_pmt_ind]][cur_acct_num[:]], columns = ['PI_PMT'])
    pi_pmt_index = list(pi_pmt.index)

    # Note rate for each account
    for i in range(0, len(col_ind3)):
        if "NOTE_RATE" in col_ind3[i]:
            note_rate_ind = i
    note_rate = pd.DataFrame(df3[col_ind3[note_rate_ind]][cur_acct_num[:]], columns = ['NOTE_RATE'])
    note_rate_index = list(note_rate.index)

    try:
        # Service fee for each account
        for i in range(0, len(col_ind3)):
            if "BSI SERVICE FEE" in col_ind3[i]:
                svcr_fee_ind = i
        
        # In Month 1707 there is no service fee      
        if is_empty(svcr_fee_ind) == False:        
            svcr_fee = pd.DataFrame(df3[col_ind3[svcr_fee_ind]][cur_acct_num[:]])
            svcr_fee_index = list(svcr_fee.index)    
        else:
            svcr_fee = pd.DataFrame({'BSI SERVICE FEE', end_bal_index })
    except:
        pass
    
    # Deferred balances for each account
    for i in range(0, len(col_ind3)):
        if "DEFERRED_BAL" in col_ind3[i]:
            defer_bal_ind = i
    defer_bal = pd.DataFrame(df3[col_ind3[defer_bal_ind]][cur_acct_num[:]], columns = ['DEFERRED_BAL'])
    defer_bal_index = list(defer_bal.index)

    ###########################################################################
    # Search for COLLECTION - Z510 worksheet                                  #
    ###########################################################################
    for i in range(0, len(ws_names)):
        if "COLLECTION - Z510" in ws_names[i]:
            col_z = i
    # COLLECTION - Z510 worksheet        
    global df4, acct_num_coll, df5
    df4 = pd.read_excel(cur_mon_remit, sheetname = ws_names[col_z])
    col_ind4 = list(df4.columns)
    # We need Column: REG_INT_AND_LIQ_INT, PRIN_COLL, ADDITIONAL_PRIN, OTHER_FEES
    df4 = pd.pivot_table(df4, index=col_ind4[3], values = ['ADDITIONAL_PRIN', 'OTHER_FEES', 'PRIN_COLL', 'REG_INT_AND_LIQ_INT'], aggfunc=np.sum)
    acct_num_coll = list(df4.index)

    ###########################################################################
    # Search for ADVANCE RECOVERY worksheet                                   #
    ###########################################################################
    for i in range(0, len(ws_names)):
        if "ADVANCE RECOVERY" in ws_names[i]:
            adv_rec = i
    df5 = pd.read_excel(cur_mon_remit, sheetname = ws_names[adv_rec])
    col_ind5 = list(df5.columns)
    for i in range(0, len(col_ind5)):
        if "ACCOUNT_NUMBER" in col_ind5[i]:
            acct_num_ind = i
            df5 = pd.read_excel(cur_mon_remit, sheetname = ws_names[adv_rec], index_col = acct_num_ind)
        else:
            pass
    df5 = df5.pivot_table(df5, index = col_ind5[3], aggfunc=np.sum)
    df5 = pd.DataFrame(df5['AMOUNT'][:], columns = ['AMOUNT'])
    row_ind5 = list(df5.index)
    
#########################################################################################################
# Modify current month remit template fiel                                                              #
# This block will place in the beg bal and other fields                                                 #
#########################################################################################################
def write_temp():
    # Load the new current month file
    wb = load_workbook(cur_mon_file)
    ###########################################################################
    # Worksheet 1, BSI raw data input                                         #
    ###########################################################################
    ws = wb['BSI']
    
    # Posting beginning balance from ending balance from last month
    for i in range(0, len(acct_num)):
        ws['B' + str(2 + i)] = beg_bal[i]
        
    # Posting end_bal
    for i in range(0, len(acct_num)):
        ws['C' + str(2 + i)].value = 0
        for j in range(0, len(end_bal)):
            if (ws['A' + str(2 + i)].value == int(end_bal_index[j])):
                ws['C' + str(2 + i)] = end_bal['END_UPB'][end_bal_index[j]]
             
    # Posting pi_pmt
    for i in range(0, len(acct_num)):
        ws['D' + str(2 + i)].value = 0
        for j in range(0, len(pi_pmt)):
            if (ws['A' + str(2 + i)].value == int(pi_pmt_index[j])):
                ws['D' + str(2 + i)] = pi_pmt['PI_PMT'][pi_pmt_index[j]]
                
    # Posting note_rate
    for i in range(0, len(acct_num)):
        ws['E' + str(2 + i)].value = 0
        for j in range(0, len(note_rate)):
            if (ws['A' + str(2 + i)].value == int(note_rate_index[j])):
                ws['E' + str(2 + i)] = note_rate['NOTE_RATE'][note_rate_index[j]]
        
    # Posting svcr_fee
    for i in range(0, len(acct_num)):
        ws['G' + str(2 + i)].value = 0
        for j in range(0, len(svcr_fee)):
            if (ws['A' + str(2 + i)].value == int(svcr_fee_index[j])):
                ws['G' + str(2 + i)] = svcr_fee['BSI SERVICE FEE'][svcr_fee_index[j]]      
    
    # Posting defer_bal
    for i in range(0, len(acct_num)):
        ws['M' + str(2 + i)].value = 0
        for j in range(0, len(defer_bal)):
            if (ws['A' + str(2 + i)].value == int(defer_bal_index[j])):
                ws['M' + str(2 + i)] = defer_bal['DEFERRED_BAL'][defer_bal_index[j]]

    # Posting gross interest
    for i in range(0, len(acct_num)):
        ws['F' + str(2 + i)].value = 0
        for j in range(0, len(acct_num_coll)):
            if (ws['A' + str(2 + i)].value == int(acct_num_coll[j])):
                ws['F' + str(2 + i)] = df4['REG_INT_AND_LIQ_INT'][acct_num_coll[j]]

    # Posting PRIN_COLL
    for i in range(0, len(acct_num)):
        ws['H' + str(2 + i)].value = 0
        for j in range(0, len(acct_num_coll)):
            if (ws['A' + str(2 + i)].value == int(acct_num_coll[j])):
                ws['H' + str(2 + i)] = df4['PRIN_COLL'][acct_num_coll[j]]

    # Posting ADDITIONAL_PRIN
    for i in range(0, len(acct_num)):
        ws['I' + str(2 + i)].value = 0
        for j in range(0, len(acct_num_coll)):
            if (ws['A' + str(2 + i)].value == int(acct_num_coll[j])):
                ws['I' + str(2 + i)] = df4['ADDITIONAL_PRIN'][acct_num_coll[j]]

    # Posting gross interest
    for i in range(0, len(acct_num)):
        ws['K' + str(2 + i)].value = 0
        for j in range(0, len(acct_num_coll)):
            if (ws['A' + str(2 + i)].value == int(acct_num_coll[j])):
                ws['K' + str(2 + i)] = df4['OTHER_FEES'][acct_num_coll[j]]

    # Posting gross interest
    for i in range(0, len(acct_num)):
        ws['J' + str(2 + i)].value = 0
        for j in range(0, len(row_ind5)):
            if (ws['A' + str(2 + i)].value == int(row_ind5[j])):
                ws['J' + str(2 + i)] = df5['AMOUNT'][row_ind5[j]]

    ###########################################################################
    # Worksheet 2 ,Remittance Report Data for                                 #
    ###########################################################################
    ws2 = wb['Remittance Report']
    
    # Find the Corp Recoveries Cell in column F(Column 6)
    for i in range(11, 100):
        if 'Corp Recoveries' in str(ws2['F' + str(i)].value):
            corp_rec_cell = ws2['F' + str(i)].value
            if is_empty(corp_rec_cell) == False:
                row = i
                ws2['H' + str(row)].value = corp_rec
                break
    
    # Find the 'Servicing fee-Previous Servicer' cell 
    for i in range(11, 100):
        if 'Servicing fee-Previous Servicer' in str(ws2['E' + str(i)].value):
            svcr_fee_prev_fee = ws2['E' + str(i)].value
            if is_empty(svcr_fee_prev_fee) == False:
                row = i
                ws2['G' + str(row)].value = prev_svcr_fee
                break
            
    # Find the first special cell adjustment        
    for i in range(11, 100):
        if 'special' in str(ws2['I' + str(i)].value):
            spec1 = ws2['I' + str(i)].value
            if is_empty(spec1) == False:
                row = i
                ws2['I' + str(row + 1)].value = special_cell
                break        
            
    try:
        for i in range(11, 100):
            if 'SPECIAL' in str(ws2['M' + str(i)].value):
                spec2 = ws2['M' + str(i)].value
                if is_empty(spec2) == False:
                    if is_empty(special_cell2) == False:
                        row = i
                        ws2['N' + str(row)].value = special_cell2
                        break 
                    else:
                        row = i
                        ws2['N' + str(row)].value = 0
                        break
    except:
        pass
    wb.save(cur_mon_file)


#########################################################################################################
# This block will be used later to combine the 3 servicers together                                     #
#########################################################################################################
#def combo():
#    sys.exit()
    
if __name__ == '__main__':
    copy_temp()
    read_remit()
    write_temp()
    sys.exit()